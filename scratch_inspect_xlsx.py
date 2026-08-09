import os
import openpyxl
import pandas as pd
import sys

# Reconfigure stdout to use utf-8
sys.stdout.reconfigure(encoding='utf-8')

workspace_dir = "c:\\Users\\USER\\OneDrive\\Documents\\universal-market-app"
excel_files = [f for f in os.listdir(workspace_dir) if f.endswith(".xlsx") and not f.startswith("~$")]

print(f"Found {len(excel_files)} Excel files:")
for idx, f in enumerate(sorted(excel_files)):
    fpath = os.path.join(workspace_dir, f)
    try:
        wb = openpyxl.load_workbook(fpath, read_only=True)
        sheets = wb.sheetnames
        print(f"\n[{idx+1}] {f} - Sheets: {sheets}")
        
        # Let's inspect the first sheet's columns
        if sheets:
            sheet_name = sheets[0]
            sh = wb[sheet_name]
            # Read first row
            rows_iter = sh.iter_rows(values_only=True)
            try:
                headers = next(rows_iter)
                print(f"  First sheet headers: {headers[:10]}")
                # Read next row
                next_row = next(rows_iter)
                print(f"  Sample row: {next_row[:10]}")
            except StopIteration:
                print("  Sheet is empty.")
    except Exception as e:
        print(f"Error reading {f}: {e}")
