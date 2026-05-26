import os
import sys
import openpyxl
import pandas as pd
import numpy as np

# Ensure UTF-8 output
sys.stdout.reconfigure(encoding='utf-8')

class StrategyHelper:
    def __init__(self, workspace_dir="c:/Users/USER/OneDrive/Documents/universal-market-app"):
        self.workspace_dir = workspace_dir
        self.strategies = {
            "Discount Stock Strategy v2 (DSS2)": {
                "file": "DSS2BEST.xlsx",
                "sheet": "Sheet1",
                "type": "trade_log",
                "date_col": "Exit Date",
                "profit_col": "Profit",
                "equity_col": "Portfolio Equity",
                "initial_capital": 100000.0,
                "desc": "Buys undervalued Indian stocks based on Graham Intrinsic Value when a UT Bot trend is confirmed."
            },
            "Chess Trading Strategy": {
                "file": "chess_battle_results.xlsx",
                "sheet": "All Trades",
                "type": "trade_log",
                "date_col": "Exit Date",
                "profit_col": "Profit",
                "equity_col": "Equity",
                "initial_capital": 100000.0,
                "desc": "A tactical system inspired by Chess positional hedging (Knights) and momentum breakout expansions (Pawns)."
            },
            "Commodity Arbitrage": {
                "file": "commodity_arb_results_TRIAL3.xlsx",
                "sheets": ["Precious_Metals", "Energy_Complex", "Agriculture_Grains", "Industrial_Base_Metals"],
                "type": "daily_pnl",
                "initial_capital": 100000.0,
                "desc": "Statistical arbitrage trading highly correlated commodity pairs using spread Z-score mean reversion."
            },
            "Crypto Arbitrage": {
                "file": "crypto_arb_results.xlsx",
                "sheets": ["Crypto_Core", "Platform_Giants", "Layer_1_Alternative", "Layer_1_Dominance"],
                "type": "daily_pnl",
                "initial_capital": 100000.0,
                "desc": "Intraday statistical arbitrage trading cryptocurrency token spreads based on statistical boundaries."
            },
            "HFT Vector Bundle": {
                "file": "hft_vector_bundle_results.xlsx",
                "sheet": "HFT Trades",
                "type": "trade_log",
                "date_col": "Exit Date",
                "profit_col": "Profit",
                "initial_capital": 100000.0,
                "desc": "Intraday high-frequency predictive models exploiting order book flow imbalance via machine learning."
            },
            "Market Geometry": {
                "file": "market_geometry_results.xlsx",
                "sheet": "Trades",
                "type": "trade_log",
                "date_col": "Exit Date",
                "profit_col": "Profit",
                "initial_capital": 100000.0,
                "desc": "Technical charting strategy executing trades on Fibonacci, Andrews Pitchfork, and geometric trend channel bounds."
            },
            "Basket Selection Strategy (BSS)": {
                "file": "trade_log_gl.xlsx",
                "sheet": "All Trades",
                "type": "trade_log",
                "date_col": "Exit Date",
                "profit_col": "Profit",
                "equity_col": "Equity",
                "initial_capital": 100000.0,
                "desc": "Selects dynamic baskets of high-momentum Indian equities, managed with ATR trailing stop-losses."
            },
            "Vector HFT": {
                "file": "vector_hft_results.xlsx",
                "sheet": "VECTOR",
                "type": "trade_log",
                "date_col": "Exit Date",
                "profit_col": "Profit",
                "initial_capital": 100000.0,
                "desc": "High-performance vectorized momentum and trend-following model using microstructural features."
            }
        }
        self.cached_curves = {}
        self.cached_trades = {}

    def get_file_path(self, filename):
        return os.path.join(self.workspace_dir, filename)

    def load_strategy_trades(self, name):
        """Loads the trade ledger for a specific strategy."""
        if name in self.cached_trades:
            return self.cached_trades[name]
        
        cfg = self.strategies.get(name)
        if not cfg:
            return pd.DataFrame()
            
        fpath = self.get_file_path(cfg["file"])
        if not os.path.exists(fpath):
            return pd.DataFrame()
            
        try:
            if cfg["type"] == "trade_log":
                wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
                sheet_name = cfg["sheet"]
                if sheet_name not in wb.sheetnames:
                    return pd.DataFrame()
                sh = wb[sheet_name]
                rows_iter = sh.iter_rows(values_only=True)
                cols = next(rows_iter)
                rows = list(rows_iter)
                df = pd.DataFrame(rows, columns=cols)
                
                # Dynamic column mapping
                # 1) Date column
                actual_date_col = None
                for c in df.columns:
                    if str(c).strip().lower() in ["exit date", "date", "timestamp", "exit_date"]:
                        actual_date_col = c
                        break
                if not actual_date_col:
                    actual_date_col = cfg["date_col"] if cfg["date_col"] in df.columns else df.columns[0]
                
                df["Exit Date"] = pd.to_datetime(df[actual_date_col], errors='coerce')
                
                # Sort by Date
                df = df.dropna(subset=["Exit Date"]).sort_values("Exit Date").reset_index(drop=True)
                
                # 2) Profit column
                actual_profit_col = None
                for c in df.columns:
                    if any(term in str(c).strip().lower() for term in ["profit", "pnl", "p&l", "yield", "gain"]):
                        actual_profit_col = c
                        break
                if not actual_profit_col:
                    actual_profit_col = cfg["profit_col"] if cfg["profit_col"] in df.columns else df.columns[0]
                    
                df["Profit"] = pd.to_numeric(df[actual_profit_col], errors='coerce').fillna(0.0)
                
                # 3) Return column
                actual_ret_col = None
                for c in df.columns:
                    if "return" in str(c).strip().lower():
                        actual_ret_col = c
                        break
                if actual_ret_col:
                    df["Return %"] = pd.to_numeric(df[actual_ret_col], errors='coerce').fillna(0.0)
                else:
                    df["Return %"] = 0.0
                    
                # 4) Stock/Asset column
                actual_stock_col = None
                for c in df.columns:
                    if str(c).strip().lower() in ["stock", "symbol", "ticker", "asset", "pair"]:
                        actual_stock_col = c
                        break
                if actual_stock_col:
                    df["Stock"] = df[actual_stock_col]
                else:
                    df["Stock"] = "Generic Asset"
                    
                # 5) Exit Reason column
                actual_reason_col = None
                for c in df.columns:
                    if "reason" in str(c).strip().lower() or "mode" in str(c).strip().lower() or "type" in str(c).strip().lower():
                        actual_reason_col = c
                        break
                if actual_reason_col:
                    df["Exit Reason"] = df[actual_reason_col]
                else:
                    df["Exit Reason"] = "Normal Exit"
                
                # Clean other numeric columns if present
                for c in df.columns:
                    if "price" in str(c).strip().lower() or "invested" in str(c).strip().lower():
                        df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0.0)
                
                self.cached_trades[name] = df
                return df
                
            elif cfg["type"] == "daily_pnl":
                # For arbitrage, we can build a trade log from PnL adjustments in sub-sheets if needed,
                # but let's load from Master_Trade_Ledger sheet if available
                wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
                sheet_name = "Master_Trade_Ledger"
                if sheet_name in wb.sheetnames:
                    sh = wb[sheet_name]
                    rows_iter = sh.iter_rows(values_only=True)
                    cols = next(rows_iter)
                    rows = list(rows_iter)
                    df = pd.DataFrame(rows, columns=cols)
                    
                    # Create generic date from row indices since raw ledger does not have dates
                    # Spread them across the backtest range (2016-2026)
                    dates = pd.date_range(start="2016-06-02", end="2026-05-04", periods=len(df))
                    df["Exit Date"] = dates
                    df["Entry Date"] = dates - pd.Timedelta(days=5)
                    
                    p_col = None
                    for c in df.columns:
                        if any(term in str(c).strip().lower() for term in ["pnl", "profit", "yield", "gain"]):
                            p_col = c
                            break
                    if not p_col:
                        p_col = cols[2]
                        
                    df["Profit"] = pd.to_numeric(df[p_col], errors='coerce').fillna(0.0)
                    df["Return %"] = (df["Profit"] / 100000.0) * 100.0
                    df["Stock"] = df["Pair"] if "Pair" in df.columns else ("Asset" if "Asset" in df.columns else "Spread")
                    df["Exit Reason"] = df["Reason"] if "Reason" in df.columns else ("Type" if "Type" in df.columns else "Mean Reversion")
                    
                    self.cached_trades[name] = df
                    return df
                else:
                    # Construct simple trade log from daily cumulative PnL jumps
                    daily_curve = self.load_daily_equity_curve(name)
                    df = pd.DataFrame()
                    df["Exit Date"] = daily_curve["Date"]
                    df["Entry Date"] = daily_curve["Date"]
                    df["Profit"] = daily_curve["Equity"].diff().fillna(0.0)
                    df["Return %"] = daily_curve["Equity"].pct_change().fillna(0.0) * 100.0
                    df["Stock"] = "ARBITRAGE"
                    df["Exit Reason"] = "Daily Rebalance"
                    df = df[df["Profit"] != 0].reset_index(drop=True)
                    self.cached_trades[name] = df
                    return df
        except Exception as e:
            print(f"Error loading trades for {name}: {e}")
            return pd.DataFrame()

    def load_daily_equity_curve(self, name):
        """Generates a continuous daily equity series for a specific strategy."""
        if name in self.cached_curves:
            return self.cached_curves[name]
            
        cfg = self.strategies.get(name)
        if not cfg:
            return pd.DataFrame()
            
        fpath = self.get_file_path(cfg["file"])
        if not os.path.exists(fpath):
            return pd.DataFrame()
            
        try:
            if cfg["type"] == "trade_log":
                trades = self.load_strategy_trades(name)
                if trades.empty:
                    return pd.DataFrame()
                
                # Check if spreadsheet has pre-calculated Equity (e.g. Portfolio Equity, Equity, etc.)
                actual_eq_col = None
                for c in trades.columns:
                    if "equity" in str(c).strip().lower() or "capital" in str(c).strip().lower():
                        actual_eq_col = c
                        break
                
                if actual_eq_col:
                    eq_series = trades[["Exit Date", actual_eq_col]].copy()
                    eq_series.columns = ["Date", "Equity"]
                else:
                    # Reconstruct from profit
                    trades_sorted = trades.sort_values("Exit Date").reset_index(drop=True)
                    trades_sorted["Cumulative Profit"] = trades_sorted["Profit"].cumsum()
                    trades_sorted["Equity"] = cfg["initial_capital"] + trades_sorted["Cumulative Profit"]
                    eq_series = trades_sorted[["Exit Date", "Equity"]].copy()
                    eq_series.columns = ["Date", "Equity"]
                
                # Ensure date is date type and drop duplicates
                eq_series["Date"] = pd.to_datetime(eq_series["Date"]).dt.normalize()
                # Aggregate to end of day if multiple trades exist
                eq_series = eq_series.groupby("Date")["Equity"].last().reset_index()
                
                # Re-index to standard business days
                start_date = max(eq_series["Date"].min(), pd.to_datetime("2016-06-02"))
                end_date = min(eq_series["Date"].max(), pd.to_datetime("2026-05-26"))
                
                master_dates = pd.date_range(start=start_date, end=end_date, freq="B")
                master_df = pd.DataFrame({"Date": master_dates})
                
                merged = pd.merge(master_df, eq_series, on="Date", how="left")
                # Forward fill values (if no trade on that day, equity remains the same)
                merged["Equity"] = merged["Equity"].ffill().fillna(cfg["initial_capital"])
                
                self.cached_curves[name] = merged
                return merged
                
            elif cfg["type"] == "daily_pnl":
                wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
                pnl_dfs = []
                for sname in cfg["sheets"]:
                    if sname in wb.sheetnames:
                        sh = wb[sname]
                        rows_iter = sh.iter_rows(values_only=True)
                        cols = next(rows_iter)
                        rows = list(rows_iter)
                        df = pd.DataFrame(rows, columns=cols)
                        if not df.empty and "Date" in df.columns and "Cumulative_PnL" in df.columns:
                            df["Date"] = pd.to_datetime(df["Date"]).dt.normalize()
                            df["Cumulative_PnL"] = pd.to_numeric(df["Cumulative_PnL"], errors='coerce').fillna(0.0)
                            pnl_dfs.append(df[["Date", "Cumulative_PnL"]].rename(columns={"Cumulative_PnL": sname}))
                
                if not pnl_dfs:
                    return pd.DataFrame()
                    
                # Merge all sub-sheets on Date
                merged = pnl_dfs[0]
                for other in pnl_dfs[1:]:
                    merged = pd.merge(merged, other, on="Date", how="outer")
                
                # Sort and interpolate
                merged = merged.sort_values("Date").reset_index(drop=True)
                merged = merged.ffill().fillna(0.0)
                
                # Calculate total daily PnL
                merged["Total_PnL"] = merged[cfg["sheets"]].sum(axis=1)
                
                # Add to initial capital
                merged["Equity"] = cfg["initial_capital"] + merged["Total_PnL"]
                
                res = merged[["Date", "Equity"]].copy()
                
                # Resample to daily Business Days
                start_date = max(res["Date"].min(), pd.to_datetime("2016-06-02"))
                end_date = min(res["Date"].max(), pd.to_datetime("2026-05-26"))
                
                master_dates = pd.date_range(start=start_date, end=end_date, freq="B")
                master_df = pd.DataFrame({"Date": master_dates})
                
                final_res = pd.merge(master_df, res, on="Date", how="left").ffill().fillna(cfg["initial_capital"])
                
                self.cached_curves[name] = final_res
                return final_res
        except Exception as e:
            print(f"Error loading daily curve for {name}: {e}")
            return pd.DataFrame()

    def get_aligned_strategy_returns(self, selected_strategies):
        """Aligns selected strategy daily returns into a single DataFrame."""
        aligned_df = None
        for name in selected_strategies:
            curve = self.load_daily_equity_curve(name)
            if curve.empty:
                continue
            
            # Compute daily return %
            curve["Daily_Return"] = curve["Equity"].pct_change().fillna(0.0)
            strat_df = curve[["Date", "Equity", "Daily_Return"]].copy()
            strat_df.columns = ["Date", f"{name}_Equity", f"{name}_Return"]
            
            if aligned_df is None:
                aligned_df = strat_df
            else:
                aligned_df = pd.merge(aligned_df, strat_df, on="Date", how="outer")
                
        if aligned_df is not None:
            aligned_df = aligned_df.sort_values("Date").reset_index(drop=True)
            # Fill NaNs by forward filling equity and setting returns to 0
            for name in selected_strategies:
                eq_col = f"{name}_Equity"
                ret_col = f"{name}_Return"
                if eq_col in aligned_df.columns:
                    cfg = self.strategies[name]
                    aligned_df[eq_col] = aligned_df[eq_col].ffill().fillna(cfg["initial_capital"])
                    aligned_df[ret_col] = aligned_df[ret_col].fillna(0.0)
        return aligned_df

    def calculate_metrics(self, df_equity, initial_capital=100000.0):
        """Calculates key trading metrics for a given equity series."""
        if df_equity.empty or len(df_equity) < 2:
            return {
                "CAGR": 0.0,
                "Sharpe": 0.0,
                "Max_DD": 0.0,
                "Win_Rate": 0.0,
                "Total_Return": 0.0,
                "Ending_Capital": initial_capital
            }
            
        equity = df_equity["Equity"].values
        dates = df_equity["Date"].values
        
        total_return = (equity[-1] / equity[0] - 1.0) * 100.0
        
        # CAGR calculation
        years = (pd.to_datetime(dates[-1]) - pd.to_datetime(dates[0])).days / 365.25
        if years > 0:
            cagr = ((equity[-1] / equity[0]) ** (1.0 / years) - 1.0) * 100.0
        else:
            cagr = 0.0
            
        # Daily returns
        daily_returns = df_equity["Equity"].pct_change().dropna()
        
        # Sharpe Ratio (annualized, assuming risk-free rate of 0% for baseline)
        if len(daily_returns) > 1 and daily_returns.std() != 0:
            sharpe = (daily_returns.mean() / daily_returns.std()) * np.sqrt(252)
        else:
            sharpe = 0.0
            
        # Max Drawdown
        peak = np.maximum.accumulate(equity)
        drawdowns = (equity - peak) / peak * 100.0
        max_dd = drawdowns.min()
        
        return {
            "CAGR": cagr,
            "Sharpe": sharpe,
            "Max_DD": max_dd,
            "Total_Return": total_return,
            "Ending_Capital": equity[-1]
        }

    def simulate_combined_portfolio(self, selected_strategies, weights, initial_capital=100000.0):
        """Simulates a portfolio combining multiple strategies based on weights."""
        if not selected_strategies or not weights:
            return pd.DataFrame()
            
        # Normalize weights
        w_sum = sum(weights)
        norm_weights = [w / w_sum for w in weights]
        
        # Load aligned daily returns
        aligned = self.get_aligned_strategy_returns(selected_strategies)
        if aligned is None or aligned.empty:
            return pd.DataFrame()
            
        # Calculate combined daily return
        combined_return = np.zeros(len(aligned))
        for name, w in zip(selected_strategies, norm_weights):
            ret_col = f"{name}_Return"
            combined_return += w * aligned[ret_col].values
            
        # Reconstruct portfolio equity curve
        portfolio_equity = np.zeros(len(aligned))
        portfolio_equity[0] = initial_capital
        for i in range(1, len(aligned)):
            portfolio_equity[i] = portfolio_equity[i-1] * (1.0 + combined_return[i])
            
        res = pd.DataFrame({
            "Date": aligned["Date"],
            "Equity": portfolio_equity,
            "Daily_Return": combined_return
        })
        return res
